"""
Recorder for historical metrics and rankings in the Binance Trading Analysis System.
Used for backtesting and machine learning purposes.
"""
import logging
import os
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime
import sqlite3
import json

# Fix relative import
import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import EXPORT_CONFIG
from ..database.manager import DatabaseManager

logger = logging.getLogger(__name__)

class MetricsRecorder:
    """Records and manages historical metrics and rankings data."""
    
    def __init__(self, db_manager: Optional[DatabaseManager] = None):
        """
        Initialize the metrics recorder.
        
        Args:
            db_manager: Optional database manager for storing results
        """
        self.db_manager = db_manager
        self.recording_id = self._generate_recording_id()
    
    def _generate_recording_id(self) -> str:
        """Generate a unique recording ID."""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f"recording_{timestamp}"
    
    def record_metrics_and_rankings(
        self,
        timeframe: str,
        metrics: Dict[str, Dict[str, Any]],
        rankings: Dict[str, Dict[str, Any]]
    ) -> bool:
        """
        Record metrics and rankings data to the database.
        
        Args:
            timeframe: Timeframe string (e.g., "1m", "1h")
            metrics: Dictionary of metrics for each symbol
            rankings: Dictionary of rankings for each symbol
            
        Returns:
            True if successful, False otherwise
        """
        if not self.db_manager:
            logger.warning("No database manager available to record metrics")
            return False
        
        try:
            # Save metrics
            self.db_manager.save_metrics(timeframe, metrics)
            
            # Save rankings
            self.db_manager.save_rankings(timeframe, rankings)
            
            logger.info(f"Recorded metrics and rankings for timeframe {timeframe}")
            return True
        except Exception as e:
            logger.error(f"Error recording metrics and rankings: {str(e)}")
            return False
    
    def record_historical_metrics_and_rankings(
        self,
        historical_data: Dict[str, List[Dict[str, Any]]],
        timeframe: str,
        start_idx: int = 0,
        end_idx: Optional[int] = None,
        step: int = 10
    ) -> bool:
        """
        Calculate and record metrics and rankings for historical data at intervals.
        
        Args:
            historical_data: Dictionary mapping symbols to lists of price data
            timeframe: Timeframe string (e.g., "1m", "1h")
            start_idx: Start index for recording
            end_idx: End index for recording (None for all data)
            step: Number of bars between recordings
            
        Returns:
            True if successful, False otherwise
        """
        from ..analysis.metrics import calculate_all_metrics
        from ..analysis.rankings import calculate_rankings
        
        if not self.db_manager:
            logger.warning("No database manager available to record historical metrics")
            return False
        
        # Determine the maximum length of historical data
        max_len = max(len(data) for data in historical_data.values()) if historical_data else 0
        
        if end_idx is None:
            end_idx = max_len - 1
        
        if start_idx >= end_idx or start_idx < 0 or end_idx >= max_len:
            logger.error(f"Invalid start/end indices: {start_idx}/{end_idx} (max: {max_len-1})")
            return False
        
        logger.info(f"Recording historical metrics for timeframe {timeframe} "
                   f"from index {start_idx} to {end_idx} with step {step}")
        
        # Process data at intervals
        for idx in range(start_idx, end_idx + 1, step):
            try:
                # Extract data up to current bar for all symbols
                current_data = {
                    symbol: data[:idx+1]
                    for symbol, data in historical_data.items()
                    if len(data) > idx
                }
                
                # Calculate metrics
                metrics = calculate_all_metrics(current_data)
                
                # Calculate rankings
                rankings = calculate_rankings(metrics)
                
                # Record metrics and rankings
                timestamp = int(datetime.now().timestamp() * 1000)
                
                # Override timestamps for historical recordings
                metrics_to_insert = []
                for symbol, metric in metrics.items():
                    # Use the timestamp from the last bar in the data
                    if symbol in current_data and current_data[symbol]:
                        bar_timestamp = current_data[symbol][-1].get('timestamp')
                        metrics_to_insert.append((
                            symbol,
                            timeframe,
                            bar_timestamp,
                            metric.get('volume_metric'),
                            metric.get('momentum_metric'),
                            metric.get('total_pct_change'),
                            metric.get('zscore_metric'),
                            metric.get('price_metric'),
                            1 if metric.get('in_uptrend', True) else 0
                        ))
                
                rankings_to_insert = []
                for symbol, ranking in rankings.items():
                    # Use the timestamp from the last bar in the data
                    if symbol in current_data and current_data[symbol]:
                        bar_timestamp = current_data[symbol][-1].get('timestamp')
                        rankings_to_insert.append((
                            symbol,
                            timeframe,
                            bar_timestamp,
                            ranking.get('volume_rank'),
                            ranking.get('momentum_rank'),
                            ranking.get('total_pct_rank'),
                            ranking.get('zscore_rank'),
                            ranking.get('price_rank'),
                            ranking.get('overall_rank'),
                            ranking.get('total_score')
                        ))
                
                # Save directly to database using bulk insert
                if metrics_to_insert:
                    self.db_manager.metrics.bulk_insert(metrics_to_insert)
                
                if rankings_to_insert:
                    self.db_manager.rankings.bulk_insert(rankings_to_insert)
                
                # Log progress every 10 steps
                if (idx - start_idx) % (step * 10) == 0:
                    logger.info(f"Processed {idx-start_idx+1}/{end_idx-start_idx+1} bars")
            
            except Exception as e:
                logger.error(f"Error processing bar {idx}: {str(e)}")
        
        logger.info(f"Completed recording historical metrics and rankings")
        return True
    
    def get_historical_metrics(
        self,
        timeframe: str,
        start_time: Optional[int] = None,
        end_time: Optional[int] = None,
        limit: int = 100
    ) -> Dict[str, List[Dict[str, Any]]]:
        """
        Retrieve historical metrics from the database.
        
        Args:
            timeframe: Timeframe string (e.g., "1m", "1h")
            start_time: Start timestamp (None for no start limit)
            end_time: End timestamp (None for no end limit)
            limit: Maximum number of records per symbol
            
        Returns:
            Dictionary mapping symbols to lists of metrics
        """
        if not self.db_manager:
            logger.warning("No database manager available to retrieve metrics")
            return {}
        
        try:
            query = """
                SELECT m.* FROM metrics m
                WHERE m.timeframe = ?
            """
            params = [timeframe]
            
            if start_time is not None:
                query += " AND m.timestamp >= ?"
                params.append(start_time)
            
            if end_time is not None:
                query += " AND m.timestamp <= ?"
                params.append(end_time)
            
            query += " ORDER BY m.timestamp DESC LIMIT ?"
            params.append(limit)
            
            # Execute query
            conn = self.db_manager.conn
            cursor = conn.cursor()
            cursor.execute(query, params)
            
            # Process results
            results = {}
            for row in cursor.fetchall():
                row_dict = dict(row)
                symbol = row_dict['symbol']
                
                if symbol not in results:
                    results[symbol] = []
                
                # Convert in_uptrend from integer to boolean
                if 'in_uptrend' in row_dict:
                    row_dict['in_uptrend'] = bool(row_dict['in_uptrend'])
                
                results[symbol].append(row_dict)
            
            return results
        
        except Exception as e:
            logger.error(f"Error retrieving historical metrics: {str(e)}")
            return {}
    
    def get_historical_rankings(
        self,
        timeframe: str,
        start_time: Optional[int] = None,
        end_time: Optional[int] = None,
        limit: int = 100
    ) -> Dict[str, List[Dict[str, Any]]]:
        """
        Retrieve historical rankings from the database.
        
        Args:
            timeframe: Timeframe string (e.g., "1m", "1h")
            start_time: Start timestamp (None for no start limit)
            end_time: End timestamp (None for no end limit)
            limit: Maximum number of records per symbol
            
        Returns:
            Dictionary mapping symbols to lists of rankings
        """
        if not self.db_manager:
            logger.warning("No database manager available to retrieve rankings")
            return {}
        
        try:
            query = """
                SELECT r.* FROM rankings r
                WHERE r.timeframe = ?
            """
            params = [timeframe]
            
            if start_time is not None:
                query += " AND r.timestamp >= ?"
                params.append(start_time)
            
            if end_time is not None:
                query += " AND r.timestamp <= ?"
                params.append(end_time)
            
            query += " ORDER BY r.timestamp DESC LIMIT ?"
            params.append(limit)
            
            # Execute query
            conn = self.db_manager.conn
            cursor = conn.cursor()
            cursor.execute(query, params)
            
            # Process results
            results = {}
            for row in cursor.fetchall():
                row_dict = dict(row)
                symbol = row_dict['symbol']
                
                if symbol not in results:
                    results[symbol] = []
                
                results[symbol].append(row_dict)
            
            return results
        
        except Exception as e:
            logger.error(f"Error retrieving historical rankings: {str(e)}")
            return {}
    
    def export_historical_data_for_ml(
        self,
        timeframe: str,
        start_time: Optional[int] = None,
        end_time: Optional[int] = None,
        export_format: str = 'csv',
        filename: Optional[str] = None
    ) -> str:
        """
        Export historical data in a format suitable for machine learning.
        
        Args:
            timeframe: Timeframe string (e.g., "1m", "1h")
            start_time: Start timestamp (None for no start limit)
            end_time: End timestamp (None for no end limit)
            export_format: Format to export (csv, xlsx, json)
            filename: Optional filename (generated if None)
            
        Returns:
            Path to the exported file
        """
        if not self.db_manager:
            logger.warning("No database manager available to export data")
            return ""
        
        # Generate filename if not provided
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"trading_data_{timeframe}_{timestamp}"
        
        # Ensure export directory exists
        os.makedirs(EXPORT_CONFIG['export_dir'], exist_ok=True)
        
        # Set file path based on format
        if export_format.lower() == 'csv':
            file_path = os.path.join(EXPORT_CONFIG['export_dir'], f"{filename}.csv")
        elif export_format.lower() == 'xlsx':
            file_path = os.path.join(EXPORT_CONFIG['export_dir'], f"{filename}.xlsx")
        elif export_format.lower() == 'json':
            file_path = os.path.join(EXPORT_CONFIG['export_dir'], f"{filename}.json")
        else:
            logger.error(f"Unsupported export format: {export_format}")
            return ""
        
        try:
            # Get data for machine learning
            data = self.db_manager.get_data_for_machine_learning(timeframe, start_time, end_time)
            
            if not data:
                logger.warning("No data available for export")
                return ""
            
            # Export based on format
            if export_format.lower() == 'csv':
                self._export_to_csv(data, file_path)
            elif export_format.lower() == 'xlsx':
                self._export_to_xlsx(data, file_path)
            elif export_format.lower() == 'json':
                self._export_to_json(data, file_path)
            
            logger.info(f"Exported data to {file_path}")
            return file_path
        
        except Exception as e:
            logger.error(f"Error exporting historical data: {str(e)}")
            return ""
    
    def _export_to_csv(self, data: List[Dict[str, Any]], file_path: str) -> None:
        """
        Export data to CSV format.
        
        Args:
            data: List of data dictionaries
            file_path: Path to save the CSV file
        """
        import csv
        
        # Get column names from first row
        if not data:
            logger.warning("No data to export to CSV")
            return
        
        columns = list(dict(data[0]).keys())
        
        with open(file_path, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            
            # Write header
            writer.writerow(columns)
            
            # Write data rows
            for row in data:
                writer.writerow([row[col] for col in columns])
    
    def _export_to_xlsx(self, data: List[Dict[str, Any]], file_path: str) -> None:
        """
        Export data to Excel format.
        
        Args:
            data: List of data dictionaries
            file_path: Path to save the Excel file
        """
        try:
            import openpyxl
            from openpyxl import Workbook
            
            # Create workbook and sheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Trading Data"
            
            # Get column names from first row
            if not data:
                logger.warning("No data to export to Excel")
                return
            
            columns = list(dict(data[0]).keys())
            
            # Write header
            for col_idx, col_name in enumerate(columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)
            
            # Write data rows
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, col_name in enumerate(columns, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row_data[col_name])
            
            # Save workbook
            wb.save(file_path)
        except ImportError:
            logger.error("openpyxl is required for Excel export but is not installed")
            # Fall back to CSV
            csv_path = file_path.replace('.xlsx', '.csv')
            logger.info(f"Falling back to CSV export: {csv_path}")
            self._export_to_csv(data, csv_path)
    
    def _export_to_json(self, data: List[Dict[str, Any]], file_path: str) -> None:
        """
        Export data to JSON format.
        
        Args:
            data: List of data dictionaries
            file_path: Path to save the JSON file
        """
        # Convert data to proper dictionary format
        formatted_data = []
        for row in data:
            formatted_data.append(dict(row))
        
        with open(file_path, 'w') as jsonfile:
            json.dump(formatted_data, jsonfile, indent=2)
