"""
Data exporter for Orange3 integration with the Binance Trading Analysis System.
"""
import logging
import os
import json
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime
import sqlite3

# Fix relative import
import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import EXPORT_CONFIG

logger = logging.getLogger(__name__)

class Orange3Exporter:
    """Exports data in formats compatible with Orange3 for machine learning."""
    
    def __init__(self, db_manager=None):
        """
        Initialize the Orange3 exporter.
        
        Args:
            db_manager: Optional database manager for retrieving data
        """
        self.db_manager = db_manager
        
        # Ensure export directory exists
        os.makedirs(EXPORT_CONFIG['export_dir'], exist_ok=True)
    
    def export_backtest_results(
        self,
        backtest_id: str,
        export_format: str = 'csv',
        separate_directions: bool = True,
        filename_prefix: Optional[str] = None
    ) -> Dict[str, str]:
        """
        Export backtest results for analysis in Orange3.
        
        Args:
            backtest_id: ID of the backtest to export
            export_format: Format to export (csv, xlsx, json)
            separate_directions: Whether to export long and short trades separately
            filename_prefix: Optional prefix for filenames
            
        Returns:
            Dictionary mapping export type to file path
        """
        if not self.db_manager:
            logger.warning("No database manager available to export backtest results")
            return {}
        
        # Generate filename prefix if not provided
        if not filename_prefix:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename_prefix = f"backtest_{timestamp}"
        
        try:
            # Get backtest data
            backtest_data = self.db_manager.get_backtest_data_for_export(backtest_id)
            
            if not backtest_data or not backtest_data.get('results'):
                logger.warning(f"No backtest data found for ID: {backtest_id}")
                return {}
            
            # Prepare export paths
            export_paths = {}
            
            # Export results
            results = backtest_data['results']
            
            if separate_directions:
                # Group results by direction
                long_results = [r for r in results if r['direction'] == 'long']
                short_results = [r for r in results if r['direction'] == 'short']
                
                # Export long trades
                if long_results:
                    long_path = self._export_data(
                        long_results,
                        f"{filename_prefix}_long",
                        export_format
                    )
                    export_paths['long'] = long_path
                
                # Export short trades
                if short_results:
                    short_path = self._export_data(
                        short_results,
                        f"{filename_prefix}_short",
                        export_format
                    )
                    export_paths['short'] = short_path
            else:
                # Export all results together
                all_path = self._export_data(
                    results,
                    filename_prefix,
                    export_format
                )
                export_paths['all'] = all_path
            
            # Export summary
            if backtest_data.get('summary'):
                # Convert summary to a list of dictionaries
                summary_list = []
                
                for direction, stats in backtest_data['summary'].items():
                    if isinstance(stats, dict):
                        stats_copy = stats.copy()
                        stats_copy['direction'] = direction
                        summary_list.append(stats_copy)
                
                if summary_list:
                    summary_path = self._export_data(
                        summary_list,
                        f"{filename_prefix}_summary",
                        export_format
                    )
                    export_paths['summary'] = summary_path
            
            return export_paths
        
        except Exception as e:
            logger.error(f"Error exporting backtest results: {str(e)}")
            return {}
    
    def export_metrics_rankings_data(
        self,
        timeframe: str,
        start_time: Optional[int] = None,
        end_time: Optional[int] = None,
        export_format: str = 'csv',
        filename_prefix: Optional[str] = None
    ) -> Dict[str, str]:
        """
        Export metrics and rankings data for analysis in Orange3.
        
        Args:
            timeframe: Timeframe string (e.g., "1m", "1h")
            start_time: Start timestamp (None for no start limit)
            end_time: End timestamp (None for no end limit)
            export_format: Format to export (csv, xlsx, json)
            filename_prefix: Optional prefix for filenames
            
        Returns:
            Dictionary mapping export type to file path
        """
        if not self.db_manager:
            logger.warning("No database manager available to export metrics and rankings")
            return {}
        
        # Generate filename prefix if not provided
        if not filename_prefix:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename_prefix = f"trading_data_{timeframe}_{timestamp}"
        
        try:
            # Get combined data for machine learning
            combined_data = self.db_manager.get_data_for_machine_learning(
                timeframe, start_time, end_time
            )
            
            if not combined_data:
                logger.warning("No data available for export")
                return {}
            
            # Export combined data
            combined_path = self._export_data(
                combined_data,
                filename_prefix,
                export_format
            )
            
            return {'combined': combined_path}
        
        except Exception as e:
            logger.error(f"Error exporting metrics and rankings: {str(e)}")
            return {}
    
    def export_time_series_data(
        self,
        symbol: str,
        timeframe: str,
        metrics: List[str],
        start_time: Optional[int] = None,
        end_time: Optional[int] = None,
        export_format: str = 'csv',
        filename_prefix: Optional[str] = None
    ) -> str:
        """
        Export time series data for a specific symbol and metrics.
        
        Args:
            symbol: Trading pair symbol
            timeframe: Timeframe string (e.g., "1m", "1h")
            metrics: List of metric names to include
            start_time: Start timestamp (None for no start limit)
            end_time: End timestamp (None for no end limit)
            export_format: Format to export (csv, xlsx, json)
            filename_prefix: Optional prefix for filename
            
        Returns:
            Path to the exported file
        """
        if not self.db_manager:
            logger.warning("No database manager available to export time series data")
            return ""
        
        # Generate filename prefix if not provided
        if not filename_prefix:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename_prefix = f"{symbol}_{timeframe}_{timestamp}"
        
        try:
            # Construct SQL query
            query = """
                SELECT 
                    p.timestamp,
                    p.open,
                    p.high,
                    p.low,
                    p.close,
                    p.volume,
            """
            
            # Add requested metrics
            metric_columns = []
            for metric in metrics:
                if metric in [
                    'volume_metric', 'momentum_metric', 'total_pct_change',
                    'zscore_metric', 'price_metric', 'in_uptrend'
                ]:
                    metric_columns.append(f"m.{metric}")
            
            # Add ranking columns
            ranking_columns = []
            if 'ranks' in metrics or 'all' in metrics:
                ranking_columns = [
                    "r.volume_rank", "r.momentum_rank", "r.total_pct_rank",
                    "r.zscore_rank", "r.price_rank", "r.overall_rank"
                ]
            
            # Add all columns to the query
            all_columns = metric_columns + ranking_columns
            if all_columns:
                query += ", " + ", ".join(all_columns)
            
            query += """
                FROM price_data p
                LEFT JOIN metrics m ON p.symbol = m.symbol AND p.timestamp = m.timestamp AND p.timeframe = m.timeframe
                LEFT JOIN rankings r ON p.symbol = r.symbol AND p.timestamp = r.timestamp AND p.timeframe = r.timeframe
                WHERE p.symbol = ? AND p.timeframe = ?
            """
            
            params = [symbol, timeframe]
            
            # Add time constraints
            if start_time is not None:
                query += " AND p.timestamp >= ?"
                params.append(start_time)
            
            if end_time is not None:
                query += " AND p.timestamp <= ?"
                params.append(end_time)
            
            query += " ORDER BY p.timestamp ASC"
            
            # Execute query
            conn = self.db_manager.conn
            cursor = conn.cursor()
            cursor.execute(query, params)
            
            # Get data
            columns = [description[0] for description in cursor.description]
            data = []
            
            for row in cursor.fetchall():
                data.append(dict(zip(columns, row)))
            
            if not data:
                logger.warning(f"No time series data found for {symbol} {timeframe}")
                return ""
            
            # Export data
            file_path = self._export_data(
                data,
                filename_prefix,
                export_format
            )
            
            return file_path
        
        except Exception as e:
            logger.error(f"Error exporting time series data: {str(e)}")
            return ""
    
    def _export_data(
        self,
        data: List[Dict[str, Any]],
        filename_prefix: str,
        export_format: str
    ) -> str:
        """
        Export data to the specified format.
        
        Args:
            data: List of data dictionaries
            filename_prefix: Prefix for the filename
            export_format: Format to export (csv, xlsx, json)
            
        Returns:
            Path to the exported file
        """
        # Set file path based on format
        if export_format.lower() == 'csv':
            file_path = os.path.join(EXPORT_CONFIG['export_dir'], f"{filename_prefix}.csv")
            self._export_to_csv(data, file_path)
        elif export_format.lower() == 'xlsx':
            file_path = os.path.join(EXPORT_CONFIG['export_dir'], f"{filename_prefix}.xlsx")
            self._export_to_xlsx(data, file_path)
        elif export_format.lower() == 'json':
            file_path = os.path.join(EXPORT_CONFIG['export_dir'], f"{filename_prefix}.json")
            self._export_to_json(data, file_path)
        elif export_format.lower() == 'tab':
            file_path = os.path.join(EXPORT_CONFIG['export_dir'], f"{filename_prefix}.tab")
            self._export_to_orange_tab(data, file_path)
        else:
            logger.error(f"Unsupported export format: {export_format}")
            return ""
        
        logger.info(f"Exported data to {file_path}")
        return file_path
    
    def _export_to_csv(self, data: List[Dict[str, Any]], file_path: str) -> None:
        """
        Export data to CSV format.
        
        Args:
            data: List of data dictionaries
            file_path: Path to save the CSV file
        """
        import csv
        
        # Get column names from first row
        if not data:
            logger.warning("No data to export to CSV")
            return
        
        columns = list(data[0].keys())
        
        with open(file_path, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            
            # Write header
            writer.writerow(columns)
            
            # Write data rows
            for row in data:
                writer.writerow([row.get(col) for col in columns])
    
    def _export_to_xlsx(self, data: List[Dict[str, Any]], file_path: str) -> None:
        """
        Export data to Excel format.
        
        Args:
            data: List of data dictionaries
            file_path: Path to save the Excel file
        """
        try:
            import openpyxl
            from openpyxl import Workbook
            
            # Create workbook and sheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            
            # Get column names from first row
            if not data:
                logger.warning("No data to export to Excel")
                return
            
            columns = list(data[0].keys())
            
            # Write header
            for col_idx, col_name in enumerate(columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)
            
            # Write data rows
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, col_name in enumerate(columns, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name))
            
            # Save workbook
            wb.save(file_path)
        except ImportError:
            logger.error("openpyxl is required for Excel export but is not installed")
            # Fall back to CSV
            csv_path = file_path.replace('.xlsx', '.csv')
            logger.info(f"Falling back to CSV export: {csv_path}")
            self._export_to_csv(data, csv_path)
    
    def _export_to_json(self, data: List[Dict[str, Any]], file_path: str) -> None:
        """
        Export data to JSON format.
        
        Args:
            data: List of data dictionaries
            file_path: Path to save the JSON file
        """
        # Format data for JSON serialization
        formatted_data = []
        
        for row in data:
            formatted_row = {}
            
            for key, value in row.items():
                # Handle non-serializable types
                if isinstance(value, datetime):
                    formatted_row[key] = value.isoformat()
                else:
                    formatted_row[key] = value
            
            formatted_data.append(formatted_row)
        
        with open(file_path, 'w') as jsonfile:
            json.dump(formatted_data, jsonfile, indent=2)
    
    def _export_to_orange_tab(self, data: List[Dict[str, Any]], file_path: str) -> None:
        """
        Export data to Orange's .tab format.
        
        Args:
            data: List of data dictionaries
            file_path: Path to save the .tab file
        """
        if not data:
            logger.warning("No data to export to Orange .tab format")
            return
        
        # Get column names and determine types
        columns = list(data[0].keys())
        
        # Determine column types (continuous, discrete, string, time)
        column_types = {}
        
        for col in columns:
            # Sample values (up to 100)
            values = [row.get(col) for row in data[:100] if row.get(col) is not None]
            
            if not values:
                column_types[col] = 'string'
                continue
            
            # Check if all values are numeric
            if all(isinstance(v, (int, float)) for v in values):
                # Check if values are discrete (few unique values) or continuous
                unique_values = set(values)
                if len(unique_values) <= 5 or (len(unique_values) <= 10 and all(isinstance(v, int) for v in values)):
                    column_types[col] = 'discrete'
                else:
                    column_types[col] = 'continuous'
            # Check if all values are datetime
            elif all(isinstance(v, datetime) for v in values):
                column_types[col] = 'time'
            # Default to string
            else:
                column_types[col] = 'string'
        
        # Mark target variables
        target_columns = [
            'pnl', 'exitReason', 'bars_held', 'winRate', 'totalTrades', 
            'winningTrades', 'losingTrades', 'averagePnl', 'totalPnl'
        ]
        
        # Write to .tab file
        with open(file_path, 'w') as f:
            # Write variable names
            f.write('\t'.join(columns) + '\n')
            
            # Write variable types
            type_line = []
            for col in columns:
                type_str = column_types[col]
                
                # Mark target variables with 'class' for classification or 'target' for regression
                if col in target_columns:
                    if type_str == 'discrete':
                        type_str = 'class'
                    elif type_str == 'continuous':
                        type_str = 'target'
                
                type_line.append(type_str)
            
            f.write('\t'.join(type_line) + '\n')
            
            # Write empty line for meta-attributes (not used)
            f.write('\t' * (len(columns) - 1) + '\n')
            
            # Write data
            for row in data:
                values = []
                
                for col in columns:
                    value = row.get(col)
                    
                    # Format the value based on its type
                    if value is None:
                        values.append('?')  # Orange's missing value symbol
                    elif isinstance(value, datetime):
                        values.append(value.isoformat())
                    elif isinstance(value, bool):
                        values.append(str(value).lower())
                    else:
                        values.append(str(value))
                
                f.write('\t'.join(values) + '\n')
